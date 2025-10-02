import logging
from typing import Optional, Dict, List, Tuple, Any

import psycopg
from psycopg.types.json import Jsonb
import openpyxl

import configs


def postgres_init():
    """
    Инициализирует соединение с PostgreSQL базой данных.

    Функция пытается подключиться к базе данных, используя конфигурационные данные из `configs.sql_database`.
    :return: Кортеж из курсора и соединения, если подключение успешно.
    None: Если возникла ошибка подключения, выводится сообщение об ошибке и возвращается None.
    """

    try:
        # Подключение к существующей базе данных
        conn = psycopg.connect(
            f"dbname={configs.sql_database['database']} "
            f"user={configs.sql_database['user']} "
            f"password={configs.sql_database['password']} "
            f"host={configs.sql_database['host']} "
            f"port={configs.sql_database['port']}"
        )
        cursor = conn.cursor()
        return conn, cursor
    except psycopg.Error as error:  # Ловим только ошибки PostgreSQL
        logging.error(f"Ошибка подключения к PostgreSQL: {error}")
        return None


def insert_user_data_in_bd(user_id: int, user_data: Dict[str, Any]) -> None:
    conn, cursor = postgres_init()
    try:
        cursor.execute('''
            INSERT INTO user_data 
            (user_id, user_data, is_admin, opened_cases, title) 
            VALUES (%s, %s, %s, %s, %s) 
            ON CONFLICT (user_id) 
            DO UPDATE SET user_data = EXCLUDED.user_data;
        ''', (user_id, Jsonb(user_data), False, 0, 'Новичок'))
        conn.commit()
    finally:
        conn.close()


def parse_and_save_to_db(file_path: str, message, bot) -> None:
    """
    Парсит данные из Excel файла и сохраняет их в базу данных PostgreSQL.

    Ожидаемые колонки: number, name, color, effect, description, collection

    Args:
        file_path: Путь к Excel файлу
        message: Объект сообщения Telegram
        bot: Объект бота Telegram
    """
    try:
        logging.debug(f"Начало обработки файла: {file_path}")

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        conn, cursor = postgres_init()
        logging.debug("Подключение к БД установлено")

        cursor.execute("DELETE FROM coupons")
        conn.commit()
        logging.debug("Старые данные удалены")

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Берем только первые 6 значений и проверяем, что они не None
                filtered_row = [str(cell).strip() if cell is not None else "" for cell in row[:6]]

                if len(filtered_row) < 6:
                    raise ValueError(
                        f"Недостаточно данных, получено только {len(filtered_row)} значений")

                number, name, color, effect, description, collection = filtered_row
                logging.debug(f"Обработка строки {row_num}: {filtered_row}")

                if not number.isdigit():
                    raise ValueError(f"Номер должен быть числом, получено: {number}")

                id = f'{collection}_{color}_{number}'

                cursor.execute(
                    "INSERT INTO coupons (id, number, name, color, effect, description, collection) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (id, int(number), name, color, effect, description, collection)
                )
                conn.commit()

            except ValueError as ve:
                logging.warning(f"Пропуск строки {row_num}: {ve}")
                continue
            except Exception as e:
                logging.error(f"Ошибка в строке {row_num}: {e}", exc_info=True)
                continue

        conn.close()
        logging.info(f"Успешно обработано {sheet.max_row - 1} строк")
        bot.reply_to(message, "Данные успешно загружены в базу данных!")

    except Exception as e:
        logging.critical(f"Ошибка обработки файла: {e}", exc_info=True)
        bot.reply_to(message, f"Ошибка: {str(e)}")


def save_info_coupon(user_id: int, coupon_code: str, name: str, color: str) -> None:
    """
    Сохраняет информацию о купоне пользователя.

    Если у пользователя нет такого купона - создает новую запись с количеством 1.
    Если купон уже есть - увеличивает количество на 1.

    Args:
        user_id: ID пользователя
        coupon_code: Код купона

    Returns:
        None
    """
    logging.info(f"Начало обработки купона для user_id={user_id}, coupon_code={coupon_code}")

    conn, cursor = postgres_init()
    try:
        # Проверяем существование записи
        logging.debug("Проверяем наличие купона у пользователя...")
        cursor.execute(
            "SELECT quantity FROM user_coupons WHERE user_id = %s AND coupon_code = %s",
            (user_id, coupon_code)
        )
        result = cursor.fetchone()

        if result:
            # Обновляем существующую запись
            new_quantity = result[0] + 1
            logging.debug(f"Купон найден, обновляем количество на {new_quantity}")
            cursor.execute(
                "UPDATE user_coupons SET quantity = %s WHERE user_id = %s AND coupon_code = %s",
                (new_quantity, user_id, coupon_code)
            )
        else:
            # Создаем новую запись
            logging.debug("Купон не найден, создаем новую запись")
            cursor.execute(
                "INSERT INTO user_coupons (user_id, coupon_code, quantity, name, color) VALUES (%s, %s, 1, %s, %s)",
                (user_id, coupon_code, name, color)
            )

        conn.commit()
        logging.info(f"Успешно обработан купон для user_id={user_id}, coupon_code={coupon_code}")
    except Exception as e:
        logging.error(f"Ошибка при обработке купона: {str(e)}")
        raise
    finally:
        conn.close()
        logging.debug("Соединение с базой данных закрыто")
